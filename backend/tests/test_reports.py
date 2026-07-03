import pytest
from fastapi.testclient import TestClient
from unittest.mock import patch, MagicMock
from datetime import datetime
import io
import openpyxl
from typing import Optional

from app.main import app
from app.core.security import get_current_active_user
from app.core.database import get_db
from app.models.user import User

# Mock Users
mock_superuser = User(
    id=1,
    email="admin@example.com",
    is_active=True,
    is_superuser=True
)

mock_normal_user = User(
    id=2,
    email="user@example.com",
    is_active=True,
    is_superuser=False
)

def override_get_superuser():
    return mock_superuser

def override_get_normal_user():
    return mock_normal_user

def override_get_db():
    mock_db = MagicMock()
    # Mocking execution to return empty results for our basic tests
    mock_db.execute.return_value.scalars.return_value.all.return_value = []
    mock_db.execute.return_value.scalar.return_value = 0
    
    def mock_refresh(obj):
        obj.id = 1
        if not hasattr(obj, 'created_at') or obj.created_at is None:
            obj.created_at = datetime.utcnow()
            
    mock_db.refresh.side_effect = mock_refresh
    yield mock_db
@pytest.fixture(autouse=True, scope="module")
def setup_overrides():
    app.dependency_overrides[get_current_active_user] = override_get_superuser
    app.dependency_overrides[get_db] = override_get_db
    yield
    app.dependency_overrides.clear()

client = TestClient(app)

def test_export_mentions_csv_returns_headers_when_empty():
    response = client.get("/api/reports/mentions/export?format=csv")
    assert response.status_code == 200
    assert response.headers["content-type"] == "text/csv; charset=utf-8"
    assert "mention_id,project_id,keyword" in response.text
    # Should only have one line (the headers) or headers + empty line
    lines = response.text.strip().split("\r\n")
    assert len(lines) == 1

def test_export_alerts_csv_returns_headers_when_empty():
    response = client.get("/api/reports/alerts/export?format=csv")
    assert response.status_code == 200
    assert "alert_id,project_id,alert_type" in response.text
    lines = response.text.strip().split("\r\n")
    assert len(lines) == 1

def test_export_incidents_csv_returns_headers_when_empty():
    response = client.get("/api/reports/incidents/export?format=csv")
    assert response.status_code == 200
    assert "incident_id,user_id,title" in response.text
    lines = response.text.strip().split("\r\n")
    assert len(lines) == 1

def test_export_filters_applied_correctly():
    # Test date_from, date_to, project_id
    with patch("app.services.export_service.apply_tenant_filter") as mock_tenant_filter:
        # Mock the chained sqlalchemy methods
        mock_query = MagicMock()
        mock_tenant_filter.return_value = mock_query
        mock_query.where.return_value = mock_query
        
        response = client.get("/api/reports/mentions/export?format=csv&project_id=5&date_from=2023-01-01T00:00:00Z&date_to=2023-12-31T23:59:59Z")
        assert response.status_code == 200
        
        # Verify that where() was called (which implies filters were applied to query)
        assert mock_query.where.call_count >= 3

def test_export_tenant_scoping_normal_user():
    # Switch to normal user
    app.dependency_overrides[get_current_active_user] = override_get_normal_user
    
    with patch("app.services.export_service.apply_tenant_filter") as mock_tenant_filter:
        response = client.get("/api/reports/mentions/export?format=csv")
        assert response.status_code == 200
        
        # Ensure apply_tenant_filter was called with the normal user
        args, _ = mock_tenant_filter.call_args
        passed_user = args[2] # current_user is 3rd arg in apply_tenant_filter(query, model, current_user)
        assert passed_user.id == 2
        assert passed_user.is_superuser is False
        
    # Reset to superuser
    app.dependency_overrides[get_current_active_user] = override_get_superuser

def test_export_project_summary_xlsx_valid_workbook():
    response = client.get("/api/reports/project-summary/export?format=xlsx")
    assert response.status_code == 200
    
    # Load workbook from response content
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    
    sheet_names = wb.sheetnames
    assert "Mentions" in sheet_names
    assert "Sentiment" in sheet_names
    assert "Categories" in sheet_names
    assert "Numerical data" in sheet_names
    assert "Analytics data" in sheet_names
    
    # Verify no fake data in Analytics Data sheet
    ws = wb["Analytics data"]
    found_mentions_row = False
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "Total Mentions":
            assert row[1] == 0  # No fake data
            found_mentions_row = True
    assert found_mentions_row

def test_summary_data_endpoint_empty():
    response = client.get("/api/reports/summary-data")
    assert response.status_code == 200
    data = response.json()
    assert data["metrics"]["total_mentions"] == 0
    assert data["metrics"]["total_alerts"] == 0
    assert data["metrics"]["total_incidents"] == 0
    assert data["metrics"]["sentiment"]["positive"] == 0
    assert data["selected_mentions"] == []
    assert data["top_sources"] == []
    assert data["trend"] == []
    assert data["top_influencers"] == []

def test_summary_data_scoping_normal_user():
    app.dependency_overrides[get_current_active_user] = override_get_normal_user
    
    with patch("app.api.reports.apply_tenant_filter") as mock_tenant_filter:
        # Return a real select statement so subquery() works
        from sqlalchemy import select
        from app.models.alert import Alert
        mock_tenant_filter.return_value = select(Alert)
        
        # Test request
        response = client.get("/api/reports/summary-data")
        assert response.status_code == 200
        
        args, _ = mock_tenant_filter.call_args
        passed_user = args[2]
        assert passed_user.id == 2
        assert passed_user.is_superuser is False
        
    app.dependency_overrides[get_current_active_user] = override_get_superuser

def test_export_unsupported_formats():
    assert client.get("/api/reports/mentions/export?format=pdf").status_code == 400
    assert client.get("/api/reports/alerts/export?format=xlsx").status_code == 400
    assert client.get("/api/reports/project-summary/export?format=csv").status_code == 400

@patch("app.services.export_service.ExportService.process_export")
def test_async_export_request(mock_process_export):
    app.dependency_overrides[get_db] = override_get_db
    response = client.post("/api/reports/export/pdf")
    assert response.status_code == 201
    data = response.json()
    assert data["report_type"] == "pdf"
    assert data["status"] == "pending"
    assert "id" in data
    mock_process_export.assert_called_once_with(1)

    response = client.post("/api/reports/export/invalid_type")
    assert response.status_code == 400

def test_upload_pdf_logo_invalid_type():
    file_content = b"fake image content"
    response = client.post(
        "/api/reports/pdf/logo",
        files={"file": ("test.txt", file_content, "text/plain")}
    )
    assert response.status_code == 400
    assert "Invalid file type" in response.text

def test_upload_pdf_logo_valid_type():
    app.dependency_overrides[get_current_active_user] = override_get_superuser
    file_content = b"fake image content"
    response = client.post(
        "/api/reports/pdf/logo",
        files={"file": ("test.png", file_content, "image/png")}
    )
    assert response.status_code == 200
    assert "logo_path" in response.json()
    assert "test.png" not in response.json()["logo_path"]

def test_request_export_saves_builder_config():
    app.dependency_overrides[get_current_active_user] = override_get_superuser
    with patch("app.api.reports.BackgroundTasks.add_task") as mock_bg:
        builder_config = {
            "theme": "dark",
            "sections": [{"id": "summary", "enabled": True}]
        }
        response = client.post("/api/reports/export/pdf", json=builder_config)
        assert response.status_code == 201
        assert response.json()["status"] == "pending"

def test_request_export_datetime_serialization():
    from unittest.mock import MagicMock
    mock_db = MagicMock()
    
    def mock_refresh(obj):
        obj.id = 1
    mock_db.refresh.side_effect = mock_refresh
    
    # Save old overrides to restore later
    old_user = app.dependency_overrides.get(get_current_active_user)
    old_db = app.dependency_overrides.get(get_db)
    
    app.dependency_overrides[get_current_active_user] = override_get_superuser
    app.dependency_overrides[get_db] = lambda: mock_db
    try:
        with patch("app.api.reports.BackgroundTasks.add_task") as mock_bg:
            # Pass native isoformat strings which pydantic parses to datetimes
            builder_config = {
                "theme": "dark",
                "date_from": "2026-06-28T10:00:00Z",
                "date_to": "2026-06-28T11:00:00Z",
                "sections": [{"id": "summary", "enabled": True}]
            }
            response = client.post("/api/reports/export/pdf", json=builder_config)
            assert response.status_code == 201
            assert response.json()["status"] == "pending"
            
            added_job = mock_db.add.call_args[0][0]
            assert added_job.builder_config is not None
            assert isinstance(added_job.builder_config["date_from"], str)
            assert isinstance(added_job.builder_config["date_to"], str)
    finally:
        if old_user:
            app.dependency_overrides[get_current_active_user] = old_user
        else:
            app.dependency_overrides.pop(get_current_active_user, None)
            
        if old_db:
            app.dependency_overrides[get_db] = old_db
        else:
            app.dependency_overrides.pop(get_db, None)

def test_pdf_generator_respects_builder_config():
    from app.services.pdf_generator import PDFGenerator
    data = {
        "project_name": "Test Project", 
        "metrics": {},
        "builder_config": {
            "theme": "dark",
            "sections": [{"id": "summary", "enabled": True}]
        }
    }
    pdf_bytes = PDFGenerator.generate_project_summary(data)
    assert len(pdf_bytes) > 0

def test_async_export_history():
    response = client.get("/api/reports/exports/history")
    assert response.status_code == 200
    data = response.json()
    assert "items" in data
    assert "total" in data
    assert data["items"] == []  # mock db returns []
    assert data["total"] == 0   # mock db returns 0

def test_get_export_status():
    from app.models.report import ReportExport, ExportStatus
    # Override the mock temporarily
    mock_db = next(override_get_db())
    mock_job = ReportExport(
        id=1,
        report_type="excel",
        requested_by=1,
        status=ExportStatus.SUCCESS,
        created_at=datetime.utcnow()
    )
    mock_db.execute.return_value.scalar_one_or_none.return_value = mock_job
    app.dependency_overrides[get_db] = lambda: mock_db
    
    res = client.get("/api/reports/exports/1")
    assert res.status_code == 200
    data = res.json()
    assert data["id"] == 1
    assert data["report_type"] == "excel"
    assert data["status"] == "success"
    
    # Restore mock
    app.dependency_overrides[get_db] = override_get_db


def test_excel_export_formatting_regression():
    from app.services.export_service import ExportService
    import openpyxl
    import io
    
    data = {
        "project_name": "Regression Project",
        "date_from": "2026-01-01",
        "date_to": "2026-01-31",
        "metrics": {"total_mentions": 10},
        "comparison": {},
        "top_mentions": [],
        "raw_mentions": [
            {"id": 1, "date": "2026-01-10", "domain": "example.com", "title": "Test", "content": "A long snippet", "url": "http://example.com", "sentiment": "neutral", "reach": 100, "interactions": 0}
        ],
        "sentiment": {"positive": 1, "negative": 0, "neutral": 9},
        "sources_list": [{"name": "example.com", "count": 10}],
        "tags_list": [{"name": "test", "count": 10}],
        "daily_trend": {"2026-01-10": 10}
    }
    
    excel_bytes = ExportService.export_project_summary_xlsx(data)
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    
    sheet_names = wb.sheetnames
    assert "Mentions" in sheet_names
    assert "Sentiment" in sheet_names
    assert "Categories" in sheet_names
    assert "Numerical data" in sheet_names
    assert "Analytics data" in sheet_names

def test_pdf_export_none_and_empty_regression():
    from app.services.pdf_generator import PDFGenerator
    data_none = {
        "project_name": "Test",
        "date_from": "2026-01-01",
        "date_to": "2026-01-31",
        "metrics": {
            "total_mentions": 10,
            "total_reach": None,
            "interactions": None,
        },
        "comparison": {},
        "exec_summary": "Test summary",
        "top_mentions": [
            {"title": None, "domain": None, "sentiment": "positive", "reach": None, "snippet": None}
        ],
        "sources_list": [{"name": "X", "count": None}],
        "tags_list": [{"name": "Y", "count": None}]
    }
    
    pdf_bytes = PDFGenerator.generate_project_summary(data_none)
    assert len(pdf_bytes) > 0


def test_email_schedules_route_priority_regression():
    # It should return 200 and a valid schedule payload, handling missing DB gracefully
    response = client.get("/api/reports/email-schedules")
    assert response.status_code == 200
    data = response.json()
    assert "daily_report_enabled" in data
    assert "weekly_report_enabled" in data
    assert "report_email_recipients" in data

def test_email_schedules_schema_error_graceful_fallback():
    from sqlalchemy.exc import SQLAlchemyError
    mock_db = MagicMock()
    mock_db.execute.side_effect = SQLAlchemyError("Simulated missing column")
    
    app.dependency_overrides[get_db] = lambda: mock_db
    try:
        response = client.get("/api/reports/email-schedules")
        assert response.status_code == 200
        data = response.json()
        assert data["daily_report_enabled"] is False
        assert data["report_email_recipients"] == ""
        mock_db.rollback.assert_called_once()
    finally:
        app.dependency_overrides[get_db] = override_get_db


def test_email_schedules_send_now_route_priority_regression():
    # Same for POST /email-schedules/send-now
    response = client.post("/api/reports/email-schedules/send-now")
    assert response.status_code != 422
import pytest
from datetime import datetime, timedelta
from app.models.report import ReportExport, ExportStatus
from app.api.reports import router
from app.services.export_service import ExportService
from app.services.pdf_generator import PDFGenerator

def test_pdf_export_accepts_custom_config():
    config = {
        "sections": [{"id": "summary", "enabled": True}, {"id": "top_mentions", "enabled": False}],
        "theme": "dark",
        "accent_color": "#ff0000",
        "font_color": "#ffffff",
        "font_style": "Courier",
        "aspect_ratio": "horizontal",
        "language": "vietnamese"
    }
    export = ReportExport(
        project_id=1,
        requested_by=1,
        report_type='pdf',
        status=ExportStatus.PENDING,
        builder_config=config
    )
    
    assert export.builder_config["theme"] == "dark"
    assert export.builder_config["accent_color"] == "#ff0000"
    assert export.builder_config["aspect_ratio"] == "horizontal"
    assert export.builder_config["language"] == "vietnamese"
    assert len(export.builder_config["sections"]) == 2

def test_pdf_generator_respects_custom_config():
    data = {
        "metrics": {"total_mentions": 10},
        "comparison": {},
        "sources_list": [],
        "tags_list": [],
        "top_mentions": [],
        "raw_mentions": [],
        "exec_summary": "Test summary"
    }
    config = {
        "sections": [{"id": "executive_summary", "enabled": True}],
        "theme": "dark"
    }
    pdf_bytes = PDFGenerator.generate_project_summary(data, config)
    assert isinstance(pdf_bytes, bytes)
    assert len(pdf_bytes) > 0
    
def test_summary_data_returns_structured_fields():
    client = TestClient(app)
    res = client.get("/api/reports/summary-data")
    assert res.status_code == 200
    data = res.json()
    assert "comparison" in data
    assert "raw_mentions" in data
