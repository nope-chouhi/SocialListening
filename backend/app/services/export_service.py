import csv
import io
from typing import Optional, List, Dict, Any, Generator
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, and_, func
from sqlalchemy.orm import Session

from app.models.mention import Mention, AIAnalysis
from app.models.alert import Alert
from app.models.incident import Incident
from app.models.source import Source
from app.models.keyword import KeywordGroup
from app.core.tenant import apply_tenant_filter
from app.models.user import User
from app.models.report import ReportExport, ExportStatus
from app.services.pdf_generator import PDFGenerator
from app.services.report_helpers import aggregate_period_data, calculate_period_comparison, generate_executive_summary
from datetime import timedelta
import os
import traceback
from app.core.database import SessionLocal

class ExportService:
    @staticmethod
    def _safe_str(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat()
        if hasattr(value, 'value'): # for Enums
            return str(value.value)
        return str(value).replace('\r', ' ').replace('\n', ' ')

    @staticmethod
    def export_mentions_csv(db: Session, current_user: User, filters: dict) -> Generator[str, None, None]:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            "mention_id", "project_id", "keyword", "source", "platform",
            "title", "content_excerpt", "url", "sentiment", "risk_score",
            "severity", "published_at", "created_at"
        ]
        writer.writerow(headers)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)
        
        query = apply_tenant_filter(select(Mention), Mention, current_user)
        
        if filters.get("project_id"):
            query = query.where(Mention.project_id == filters["project_id"])
        if filters.get("date_from"):
            query = query.where(Mention.published_at >= filters["date_from"])
        if filters.get("date_to"):
            query = query.where(Mention.published_at <= filters["date_to"])
            
        # Get AI Analysis map for these mentions
        # In a very large dataset, this should be joined or batched, but for this phase:
        mentions = db.execute(query).scalars().all()
        mention_ids = [m.id for m in mentions]
        
        analyses = {}
        if mention_ids:
            analyses_query = select(AIAnalysis).where(AIAnalysis.mention_id.in_(mention_ids))
            for a in db.execute(analyses_query).scalars().all():
                analyses[a.mention_id] = a
                
        # To handle sentiment and risk_level filters we apply them in-memory
        # (Could also be done with JOIN, but this is simple and matches existing patterns)
        for m in mentions:
            analysis = analyses.get(m.id)
            sentiment_str = ""
            risk_score = ""
            crisis_level = ""
            
            if analysis:
                sentiment_str = analysis.sentiment.value if hasattr(analysis.sentiment, 'value') else str(analysis.sentiment)
                risk_score = analysis.risk_score
                crisis_level = analysis.crisis_level
                
            # Apply AI filters manually if they were passed
            if filters.get("sentiment") and sentiment_str != filters["sentiment"]:
                continue
            if filters.get("risk_level"):
                # Simplistic risk level filter: low(0-30), medium(31-60), high(61-80), critical(81-100)
                rl = filters["risk_level"]
                if not analysis:
                    continue
                if rl == "low" and risk_score > 30: continue
                if rl == "medium" and (risk_score <= 30 or risk_score > 60): continue
                if rl == "high" and (risk_score <= 60 or risk_score > 80): continue
                if rl == "critical" and risk_score <= 80: continue

            excerpt = m.content[:200] + "..." if m.content and len(m.content) > 200 else m.content
            
            row = [
                m.id, m.project_id, m.keyword_text, m.source_type, m.platform,
                m.title, excerpt, m.url, sentiment_str, risk_score, crisis_level,
                m.published_at, m.collected_at
            ]
            
            writer.writerow([ExportService._safe_str(v) for v in row])
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    @staticmethod
    def export_alerts_csv(db: Session, current_user: User, filters: dict) -> Generator[str, None, None]:
        output = io.StringIO()
        writer = csv.writer(output)
        
        headers = [
            "alert_id", "project_id", "alert_type", "severity", "status",
            "title", "message", "mention_id", "created_at", "resolved_at"
        ]
        writer.writerow(headers)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)
        
        query = apply_tenant_filter(select(Alert), Alert, current_user)
        
        if filters.get("project_id"):
            query = query.where(Alert.project_id == filters["project_id"])
        if filters.get("date_from"):
            query = query.where(Alert.created_at >= filters["date_from"])
        if filters.get("date_to"):
            query = query.where(Alert.created_at <= filters["date_to"])
        if filters.get("severity"):
            query = query.where(Alert.severity == filters["severity"])
        if filters.get("status"):
            query = query.where(Alert.status == filters["status"])
            
        alerts = db.execute(query).scalars().all()
        
        for a in alerts:
            row = [
                a.id, a.project_id, "general", a.severity, a.status,
                a.title, a.message, a.mention_id, a.created_at, a.resolved_at
            ]
            writer.writerow([ExportService._safe_str(v) for v in row])
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    @staticmethod
    def export_incidents_csv(db: Session, current_user: User, filters: dict) -> Generator[str, None, None]:
        output = io.StringIO()
        writer = csv.writer(output)
        
        headers = [
            "incident_id", "user_id", "title", "status",
            "assigned_to", "created_at", "updated_at", "resolved_at"
        ]
        writer.writerow(headers)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)
        
        # apply tenant filter via user relationship for incidents if project_id isn't directly on it
        # Actually incident has user_id, mention_id, owner_id. Let's use user_id to scope
        query = select(Incident)
        if not current_user.is_superuser:
            query = query.where(Incident.user_id == current_user.id)
            
        if filters.get("date_from"):
            query = query.where(Incident.created_at >= filters["date_from"])
        if filters.get("date_to"):
            query = query.where(Incident.created_at <= filters["date_to"])
        if filters.get("status"):
            query = query.where(Incident.status == filters["status"])
            
        incidents = db.execute(query).scalars().all()
        
        for inc in incidents:
            row = [
                inc.id, inc.user_id, inc.title, inc.status,
                inc.owner_id, inc.created_at, inc.updated_at, inc.resolved_at
            ]
            writer.writerow([ExportService._safe_str(v) for v in row])
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    @staticmethod
    def export_project_summary_xlsx(export_data: dict) -> bytes:
        wb = openpyxl.Workbook()
        
        # Helper for styling headers
        def style_header(ws, color="3b82f6"):
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.freeze_panes = "A2"
            
        metrics = export_data.get('metrics', {})
        comparison = export_data.get('comparison', {})
        
        # 1. SUMMARY & ANALYTICS SHEET
        ws_summary = wb.active
        ws_summary.title = "Analytics Data"
        
        ws_summary.append(["Social Listening Project Report"])
        ws_summary["A1"].font = Font(bold=True, size=16, color="4f46e5")
        ws_summary.append(["Project:", export_data.get('project_name', 'All')])
        ws_summary.append(["Period:", f"{export_data.get('date_from')} to {export_data.get('date_to')}"])
        ws_summary.append(["Generated:", str(datetime.now())])
        ws_summary.append([])
        
        ws_summary.append(["Metric", "Current Period", "Comparison (vs Prev)"])
        style_header(ws_summary, "4f46e5")
        
        ws_summary.append(["Total Mentions", metrics.get('total_mentions', 0), comparison.get('mentions_change', '0%')])
        ws_summary.append(["Total Reach", metrics.get('total_reach', 0), comparison.get('reach_change', '0%')])
        ws_summary.append(["Interactions", metrics.get('interactions', 0), comparison.get('interactions_change', '0%')])
        
        # 2. MENTIONS SHEET
        ws_mentions = wb.create_sheet("Mentions")
        mentions_headers = ["ID", "Date", "Domain", "Title", "Content/Snippet", "URL", "Sentiment", "Reach", "Interactions"]
        ws_mentions.append(mentions_headers)
        style_header(ws_mentions, "3b82f6")
        
        for m in export_data.get('top_mentions', []): # We pass all mentions in 'top_mentions' or a separate list. Wait, in helpers I truncated top_mentions to 10. Let's pass 'all_mentions_export' from the service.
            pass # We will populate this from export_data['raw_mentions']
            
        for m in export_data.get('raw_mentions', []):
            ws_mentions.append([
                m['id'], m['date'], m['domain'], m['title'], 
                m['content'], m['url'], m['sentiment'], m['reach'], m['interactions']
            ])
            # Wrap content column
            ws_mentions.cell(row=ws_mentions.max_row, column=5).alignment = Alignment(wrap_text=True)
            
        ws_mentions.column_dimensions['D'].width = 30
        ws_mentions.column_dimensions['E'].width = 60
        ws_mentions.column_dimensions['F'].width = 30
        
        # 3. SENTIMENT SHEET
        ws_sent = wb.create_sheet("Sentiment")
        ws_sent.append(["Sentiment", "Count", "Percentage"])
        style_header(ws_sent, "10b981")
        
        total_m = metrics.get('total_mentions', 1) or 1
        for k, v in metrics.get('sentiment', {}).items():
            pct = f"{(v / total_m) * 100:.1f}%"
            ws_sent.append([k.capitalize(), v, pct])
            
        # 4. SOURCES SHEET
        ws_sources = wb.create_sheet("Sources")
        ws_sources.append(["Domain/Platform", "Mentions", "Percentage Share"])
        style_header(ws_sources, "f59e0b")
        
        for s in metrics.get('sources_list', []):
            pct = f"{(s['count'] / total_m) * 100:.1f}%"
            ws_sources.append([s['name'], s['count'], pct])
            
        # 5. NUMERICAL DATA (Daily)
        ws_num = wb.create_sheet("Numerical Data")
        ws_num.append(["Date", "Total Mentions", "Positive", "Negative", "Neutral", "Reach"])
        style_header(ws_num, "6366f1")
        
        for dt, stats in sorted(metrics.get('daily_trend', {}).items()):
            ws_num.append([dt, stats['count'], stats['positive'], stats['negative'], stats['neutral'], stats['reach']])
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def get_export_data(db: Session, current_user: User, filters: dict, builder_config: dict = None) -> dict:
        builder_config = builder_config or {}
        date_from_str = builder_config.get("date_from") or (filters.get("date_from").isoformat() if filters.get("date_from") else None)
        date_to_str = builder_config.get("date_to") or (filters.get("date_to").isoformat() if filters.get("date_to") else None)
        project_id = filters.get("project_id")
        
        mentions_query = apply_tenant_filter(select(Mention), Mention, current_user)
        project_name = "All Projects"
        if project_id:
            mentions_query = mentions_query.where(Mention.project_id == project_id)
            project = db.execute(apply_tenant_filter(select(KeywordGroup), KeywordGroup, current_user).where(KeywordGroup.id == project_id)).scalar_one_or_none()
            if project:
                project_name = project.name
                
        dt_from, dt_to = None, None
        if date_from_str:
            try: dt_from = datetime.fromisoformat(date_from_str.replace('Z', '+00:00'))
            except: pass
        if date_to_str:
            try: dt_to = datetime.fromisoformat(date_to_str.replace('Z', '+00:00'))
            except: pass
            
        if dt_from: mentions_query = mentions_query.where(Mention.published_at >= dt_from)
        if dt_to: mentions_query = mentions_query.where(Mention.published_at <= dt_to)
        
        mentions = db.execute(mentions_query).scalars().all()
        mention_ids = [m.id for m in mentions]
        analyses_list = db.execute(select(AIAnalysis).where(AIAnalysis.mention_id.in_(mention_ids))).scalars().all() if mention_ids else []
        analyses_dict = {a.mention_id: a for a in analyses_list}
        
        current_metrics = aggregate_period_data(mentions, analyses_dict)
        
        previous_metrics = {"total_mentions": 0, "total_reach": 0, "interactions": 0}
        if dt_from and dt_to:
            duration = dt_to - dt_from
            prev_from = dt_from - duration
            prev_to = dt_from
            prev_q = apply_tenant_filter(select(Mention), Mention, current_user)
            if project_id: prev_q = prev_q.where(Mention.project_id == project_id)
            prev_q = prev_q.where(and_(Mention.published_at >= prev_from, Mention.published_at < prev_to))
            prev_mentions = db.execute(prev_q).scalars().all()
            if prev_mentions:
                p_ids = [m.id for m in prev_mentions]
                p_ana = db.execute(select(AIAnalysis).where(AIAnalysis.mention_id.in_(p_ids))).scalars().all()
                previous_metrics = aggregate_period_data(prev_mentions, {a.mention_id: a for a in p_ana})
        
        comparison = calculate_period_comparison(current_metrics, previous_metrics)
        exec_summary = generate_executive_summary(current_metrics, comparison)
        
        raw_mentions = []
        for m in mentions:
            analysis = analyses_dict.get(m.id)
            sent_val = "neutral"
            if analysis:
                sent_val = analysis.sentiment.value if hasattr(analysis.sentiment, 'value') else analysis.sentiment
            inter = (m.likes_count or 0) + (m.shares_count or 0) + (m.comments_count or 0) + (m.views_count or 0)
            raw_mentions.append({
                "id": m.id,
                "date": m.published_at.strftime('%Y-%m-%d %H:%M') if m.published_at else "",
                "domain": m.domain or m.platform or "Unknown",
                "title": m.title or "",
                "content": m.snippet or m.content or "",
                "url": m.url or "",
                "sentiment": str(sent_val),
                "reach": m.reach_estimate or 0,
                "interactions": inter
            })

        return {
            "project_name": project_name,
            "date_from": date_from_str or "All time",
            "date_to": date_to_str or "All time",
            "metrics": current_metrics,
            "comparison": comparison,
            "exec_summary": exec_summary,
            "top_mentions": current_metrics.get("top_mentions", []),
            "sources_list": current_metrics.get("sources_list", []),
            "tags_list": current_metrics.get("tags_list", []),
            "raw_mentions": raw_mentions
        }

    @staticmethod
    def process_export(export_id: int):
        db = SessionLocal()
        export_job = db.execute(select(ReportExport).where(ReportExport.id == export_id)).scalar_one_or_none()
        if not export_job:
            db.close()
            return

        try:
            export_job.status = ExportStatus.RUNNING
            db.commit()

            current_user = db.execute(select(User).where(User.id == export_job.requested_by)).scalar_one_or_none()
            
            export_dir = os.path.join(os.getcwd(), 'data', 'exports')
            os.makedirs(export_dir, exist_ok=True)
            
            timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"Export_{export_id}_{timestamp_str}.{export_job.report_type}"
            file_path = os.path.join(export_dir, file_name)

            filters = {"project_id": export_job.project_id}
            builder_config = export_job.builder_config or {}
            
            export_data = ExportService.get_export_data(db, current_user, filters, builder_config)

            if export_job.report_type in ['excel', 'xlsx']:
                content = ExportService.export_project_summary_xlsx(export_data)
                with open(file_path, "wb") as f:
                    f.write(content)
            elif export_job.report_type == 'pdf':
                content = PDFGenerator.generate_project_summary(export_data)
                with open(file_path, "wb") as f:
                    f.write(content)
            else:
                raise ValueError(f"Unsupported report type: {export_job.report_type}")

            export_job.file_path = file_path
            export_job.status = ExportStatus.SUCCESS
            export_job.completed_at = datetime.now()
            db.commit()

        except Exception as e:
            export_job.status = ExportStatus.FAILED
            export_job.error_message = str(e) + "\n" + traceback.format_exc()
            export_job.completed_at = datetime.now()
            db.commit()
        finally:
            db.close()
